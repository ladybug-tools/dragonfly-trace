# coding=utf-8
"""Methods to write Dragonfly Models to Trane TRACE."""
from __future__ import division
import io
import zipfile
from collections import OrderedDict

try:
    import openpyxl
    from dragonfly_trace.exp import programs_to_exp, construction_sets_to_exp
except Exception:  # we are in Python 2 and Excel export is not possible
    openpyxl, programs_to_exp, construction_sets_to_exp = None, None, None

from ladybug.datatype.area import Area
from ladybug.datatype.distance import Distance
from ladybug.datatype.temperature import Temperature
from ladybug.datatype.rvalue import RValue
from honeybee.typing import clean_and_number_string
from dragonfly.room2d import Room2D
from dragonfly_energy.gbxml.parameters import GBXMLParameters

from .util import sort_rooms_for_trace_700
from .airflows import airflows_trace700_matrix, \
    AIRFLOW_TABLE_FORMAT, AIRFLOW_TABLE_FORMAT_62_1
from .loads import people_and_lights_trace700_matrix, \
    miscellaneous_loads_trace700_matrix, PEOPLE_AND_LIGHTS_TABLE_FORMAT, \
    MISCELLANEOUS_LOADS_TABLE_FORMAT
from .calculation import calculation_matrix


# formatting for each attribute in the rooms table
ROOM_TABLE_FORMAT = (
    'user',
    'locked',
    'locked',
    'varies_default',
    'default',
    'default',
    'user',
    'user',
    'user',
    'user',
    'user',
    'default',
    'user',
    'user',
    'varies',
    'user',
    'user',
    'default',
    'default',
    'default',
    'default',
    'default',
    'default',
    'user',
    'default',
    'default',
    'default',
    'user',
    'default'
)


def rooms_to_trace700_matrix(rooms, si_units=False):
    """Get a matrix for the "Rooms" table of the TRACE 700 Component Tree.

    Args:
        rooms: A list of dragonfly Room2Ds and honeybee Rooms for which the
            TRACE 700 "Rooms" matrix will be returned.
        si_units: Boolean to note whether the units of the values in the resulting
            matrix are in SI (True) instead of IP (False). (Default: False).

    Returns:
        A list of list where each sublist represents a row of the Rooms table
        of the TRACE 700 Component Tree.
    """
    # set up things for unit conversion
    dist_unit = 'm' if si_units else 'ft'
    r_unit = 'm2-C/W' if si_units else 'hr-ft2-F/Btu'
    temp_unit = 'C' if si_units else 'F'
    area, distance, temperature, r_value = Area(), Distance(), Temperature(), RValue()

    # set up the names of the rows
    row_names = [
        'Room Description',
        'Assigned to System',
        'Assigned to Zone',
        'Room Template',
        'Thermostat Template',
        'Construction Template',
        'Floor Length ({})'.format(dist_unit),
        'Floor Width ({})'.format(dist_unit),
        'Flr to Flr Height ({})'.format(dist_unit),
        'Plenum Height ({})'.format(dist_unit),
        'Height Above Ground ({})'.format(dist_unit),
        'Acoustic Ceiling Resistance ({})'.format(r_unit),
        'Cooling Dry Bulb ({})'.format(temp_unit),
        'Heating Dry Bulb ({})'.format(temp_unit),
        'Relative Humidity (%)',
        'Cooling Driftpoint ({})'.format(temp_unit),
        'Heating Driftpoint ({})'.format(temp_unit),
        'Thermostat Cooling Schedule',
        'Thermostat Heating Schedule',
        'Thermostat Location',
        'CO2 Sensor Location',
        'Humidity Moisture Capacitance',
        'Humidistat Location',
        'Duplicate Floor Multiplier',
        'Duplicate Rooms per Zone',
        'Room Mass / # of Hours',
        'Slab Construction Type',
        'Room Type',
        'Carpeted Floor'
    ]

    # loop through the rooms and add each of the attributes
    room_mtx = []
    for room in rooms:
        # figure out the values for certain attributes
        f2f = room.floor_to_ceiling_height if isinstance(room, Room2D) \
            else room.volume / room.floor_area
        plenum = room.ceiling_plenum_depth if isinstance(room, Room2D) else 0
        elev = room.floor_elevation if isinstance(room, Room2D) \
            else room.average_floor_height
        multiplier = room.parent.multiplier \
            if isinstance(room, Room2D) and room.has_parent else room.multiplier
        zone = room.display_name if room._zone is None else room.zone
        program = '@ {}'.format(room.properties.energy.program_type.display_name) \
            if room.properties.energy._program_type is not None else 'Default'
        c_set = '@ {}'.format(room.properties.energy.construction_set.display_name)
        set_pt = room.properties.energy.setpoint
        if room.properties.energy.setpoint is not None:
            room_type = 'Conditioned' if room.properties.energy.hvac else 'Unconditioned'
            cool_set_pt = set_pt.cooling_setpoint
            heat_set_pt = set_pt.heating_setpoint
            cool_set_back = set_pt.cooling_setback
            heat_set_back = set_pt.heating_setback
            humid_pt = set_pt.dehumidifying_setpoint \
                if set_pt.dehumidifying_setpoint is not None else '50'
        else:
            room_type = 'Unconditioned'
            cool_set_pt = 50
            heat_set_pt = 0
            cool_set_back = 50
            heat_set_back = 0
            humid_pt = '50'

        # put all attributes into a list
        room_attr = [
            room.display_name,
            'Default System',
            zone,
            program,
            'Default',
            c_set,
            room.floor_area,
            1,
            f2f,
            plenum,
            elev,
            0.31451,
            cool_set_pt,
            heat_set_pt,
            humid_pt,
            cool_set_back,
            heat_set_back,
            'None',
            'None',
            'Room',
            'None',
            'Medium',
            'Room',
            multiplier,
            1,
            'Time delay based on actual mass',
            '4" LW Concrete',
            room_type,
            'Yes'
        ]
        room_mtx.append(room_attr)

    # transpose the matrix and convert SI units to IP
    room_matrix = [list(row) for row in zip(*room_mtx)]
    if not si_units:
        room_matrix[6] = list(area.to_unit(room_matrix[6], 'ft2', 'm2'))
        room_matrix[8] = list(distance.to_unit(room_matrix[8], 'ft', 'm'))
        room_matrix[9] = list(distance.to_unit(room_matrix[9], 'ft', 'm'))
        room_matrix[10] = list(distance.to_unit(room_matrix[10], 'ft', 'm'))
        room_matrix[11] = list(r_value.to_unit(room_matrix[11], 'F-ft2-h/Btu', 'm2-K/W'))
        room_matrix[12] = list(temperature.to_unit(room_matrix[12], 'F', 'C'))
        room_matrix[13] = list(temperature.to_unit(room_matrix[13], 'F', 'C'))
        room_matrix[15] = list(temperature.to_unit(room_matrix[15], 'F', 'C'))
        room_matrix[16] = list(temperature.to_unit(room_matrix[16], 'F', 'C'))

    # round the numbers so that they display nicely
    for row_i in (6, 8, 9, 10, 11):
        room_matrix[row_i] = [round(val, 3) for val in room_matrix[row_i]]
    for row_i in (12, 13, 15, 16):
        room_matrix[row_i] = [round(val) for val in room_matrix[row_i]]

    # insert the column for the row names
    for row_name, row in zip(row_names, room_matrix):
        row.insert(0, row_name)
    return room_matrix


def model_to_trace700_matrix(
    model, si_units=False, ventilation_method='Sum of Outdoor Air',
    use_multiplier=True, exclude_plenums=True, merge_method=None, geometry_names=False
):
    """Get matrices with TRACE 700 simulation attributes of a Model.

    The resulting matrices can be written to a CSV and then copied into
    the tables that appear in the Component Tree view of TRACE 700. The
    order and organization of rooms in the resulting matrix matches that
    of the gbXML produced from the same model.

    Args:
        model: A dragonfly Model for which a TRACE 700 CSV matrix will be returned.
        si_units: Boolean to note whether the units of the values in the resulting
            matrix are in SI (True) instead of IP (False). (Default: False).
        ventilation_method: Optional text for the ventilation method to be used in the
            resulting matrix. Choose from the following.

            * Sum of Outdoor Air
            * ASHRAE 62.1

        use_multiplier: If True, the multipliers on this Model's Stories will be
            passed along to the CSV. If False, full geometry objects will be written
            for each and every floor in the building that are represented through
            multipliers and all resulting multipliers will be 1. (Default: True).
        exclude_plenums: Boolean to indicate whether ceiling/floor plenum depths
            assigned to Room2Ds should be ignored during translation. This
            results in each Room2D translating to a single Honeybee Room at
            the full floor_to_ceiling_height instead of a base Room with (a)
            plenum Room(s). (Default: True).
        merge_method: An optional text string to describe how the Room2Ds should
            be merged into individual Rooms during the translation. Specifying a
            value here can be an effective way to reduce the number of Room
            volumes in the resulting model and, ultimately, yield a faster
            simulation time in the destination engine with fewer results
            to manage. Note that Room2Ds will only be merged if they form a
            continuous volume. Otherwise, there will be multiple Rooms per
            zone or story, each with an integer added at the end of their
            identifiers. Choose from the following options:

            * None - No merging of Room2Ds will occur
            * Zones - Room2Ds in the same zone will be merged
            * PlenumZones - Only plenums in the same zone will be merged
            * Stories - Rooms in the same story will be merged
            * PlenumStories - Only plenums in the same story will be merged

        geometry_names: Boolean to note whether a cleaned version of all geometry
            display names should be used instead of identifiers when translating
            the Model to OSM and IDF. Using this flag will affect all Rooms, Faces,
            Apertures, Doors, and Shades. It will generally result in more read-able
            names in the OSM and IDF but this means that it will not be easy to map
            the EnergyPlus results back to the original Honeybee Model. Cases
            of duplicate IDs resulting from non-unique names will be resolved
            by adding integers to the ends of the new IDs that are derived from
            the name. (Default: False).

    Returns:
        A tuple with four items.

        room_matrix -- A list of list where each sublist represents a row of the
            Rooms table of the TRACE 700 Component Tree.

        airflows_matrix -- A list of list where each sublist represents a row of the
            Airflows table of the TRACE 700 Component Tree.

        people_and_lights_matrix -- A list of list where each sublist represents
            a row of the People & Lighting table of the TRACE 700 Component Tree.

        misc_loads_matrix -- A list of list where each sublist represents a row of
            the Miscellaneous Loads table of the TRACE 700 Component Tree.

        oa_calc_matrix -- A list of list where each sublist represents a row of a
            table that illustrates how the outdoor airflow rates appearing in
            the airflows_matrix were calculated.
    """
    # convert the rooms into the format in which it will go off to TRACE
    rooms_for_trace = []  # list to hold the rooms for CSV reporting
    assert len(model.room_2ds) != 0 or len(model.room_3ds) != 0, \
        'Model must have rooms to be able to export to TRACE700 CSV.'

    # scale the model if the units are not meters
    model = model.duplicate()  # duplicate model to avoid mutating it
    if model.units != 'Meters':
        model.convert_to_units('Meters')
    model.properties.energy.set_areas_by_unit_system()
    tol = model.tolerance
    # reset the IDs to be derived from the display_names if requested
    if geometry_names:
        model.reset_ids()

    # account for story multipliers, separated room plenums and room merge method
    merge_map = model._extract_merge_map(merge_method, exclude_plenums, tol)
    for building in model.buildings:
        # separate the plenums unless they are excluded
        if not exclude_plenums and building.has_room_2d_plenums:
            building.convert_plenum_depths_to_room_2ds(tol)
        # collect all of the unmerged rooms
        if use_multiplier:
            for story in building.unique_stories:
                rooms_for_trace.extend(story.room_2ds)
        else:
            for story in building.all_stories():
                rooms_for_trace.extend(story.room_2ds)

    # apply the merge map if it exists
    if merge_map is not None:
        # gather the Room objects to be merged
        merge_groups, remove_i = OrderedDict(), set()
        insert_i, insert_count = [], 0
        for i, room in enumerate(rooms_for_trace):
            try:
                merge_name = merge_map[room.identifier]
                try:
                    merge_groups[merge_name].append(room)
                except KeyError:  # first item in the group
                    merge_groups[merge_name] = [room]
                    insert_i.append(insert_count)
                remove_i.add(i)
            except KeyError:  # not a room to be merged
                insert_count += 1
        # create the new rooms and assign them to the model
        new_rooms = [r for i, r in enumerate(rooms_for_trace) if i not in remove_i]
        group_ids = {}
        zip_obj = zip(reversed(insert_i), reversed(merge_groups.items()))
        for ins_i, (group_name, room_group) in zip_obj:
            merged_rooms = Room2D.join_room_2ds(room_group, tol, tol)
            for room in merged_rooms:
                room.identifier = clean_and_number_string(group_name, group_ids)
                room.display_name = group_name
                new_rooms.insert(ins_i, room)
        rooms_for_trace = new_rooms

    # add the 3D Honeybee Rooms to the list
    for building in model.buildings:
        rooms_for_trace.extend(building.room_3ds)

    # sort the rooms so that they match their order in TRACE 700
    rooms_for_trace = sort_rooms_for_trace_700(rooms_for_trace)

    # create the matrices of data from the model rooms
    room_matrix = rooms_to_trace700_matrix(rooms_for_trace, si_units)
    airflows_matrix = airflows_trace700_matrix(rooms_for_trace, si_units, ventilation_method)
    people_and_lights_matrix = people_and_lights_trace700_matrix(rooms_for_trace, si_units)
    misc_loads_matrix = miscellaneous_loads_trace700_matrix(rooms_for_trace, si_units)

    # convert the dragonfly rooms to honeybee so that we have proper volumes for airflow
    oa_calc_matrix = None
    if ventilation_method == 'Sum of Outdoor Air':
        hb_model = model.to_honeybee(
            'District', use_multiplier=use_multiplier, exclude_plenums=exclude_plenums,
            merge_method=merge_method, solve_ceiling_adjacencies=False, enforce_adj=False
        )[0]
        hb_room_dict = {r.identifier: r for r in hb_model.rooms}
        hb_rooms_for_trace = [hb_room_dict[r.identifier] for r in rooms_for_trace]
        oa_calc_matrix = calculation_matrix(hb_rooms_for_trace, si_units)

    return room_matrix, airflows_matrix, people_and_lights_matrix, misc_loads_matrix, \
        oa_calc_matrix


def model_to_trace700_csv(
    model, si_units=False, ventilation_method='Sum of Outdoor Air',
    use_multiplier=True, exclude_plenums=True, merge_method=None, geometry_names=False
):
    """Generate a CSV string with TRACE 700 load simulation attributes of a Model.

    The resulting CSV tables can be copied into the tables that appear in the
    Component Tree view of TRACE 700. The order and organization of rooms in
    the resulting matrix should match that of the gbXML produced from the same model.

    Args:
        model: A dragonfly Model for which a TRACE 700 CSV matrix will be returned.
        si_units: Boolean to note whether the units of the values in the resulting
            matrix are in SI (True) instead of IP (False). (Default: False).
        ventilation_method: Optional text for the ventilation method to be used in the
            resulting matrix. Choose from the following.

            * Sum of Outdoor Air
            * ASHRAE 62.1

        use_multiplier: If True, the multipliers on this Model's Stories will be
            passed along to the CSV. If False, full geometry objects will be written
            for each and every floor in the building that are represented through
            multipliers and all resulting multipliers will be 1. (Default: True).
        exclude_plenums: Boolean to indicate whether ceiling/floor plenum depths
            assigned to Room2Ds should be ignored during translation. This
            results in each Room2D translating to a single Honeybee Room at
            the full floor_to_ceiling_height instead of a base Room with (a)
            plenum Room(s). (Default: True).
        merge_method: An optional text string to describe how the Room2Ds should
            be merged into individual Rooms during the translation. Specifying a
            value here can be an effective way to reduce the number of Room
            volumes in the resulting model and, ultimately, yield a faster
            simulation time in the destination engine with fewer results
            to manage. Note that Room2Ds will only be merged if they form a
            continuous volume. Otherwise, there will be multiple Rooms per
            zone or story, each with an integer added at the end of their
            identifiers. Choose from the following options:

            * None - No merging of Room2Ds will occur
            * Zones - Room2Ds in the same zone will be merged
            * PlenumZones - Only plenums in the same zone will be merged
            * Stories - Rooms in the same story will be merged
            * PlenumStories - Only plenums in the same story will be merged

        geometry_names: Boolean to note whether a cleaned version of all geometry
            display names should be used instead of identifiers when translating
            the Model to OSM and IDF. Using this flag will affect all Rooms, Faces,
            Apertures, Doors, and Shades. It will generally result in more read-able
            names in the OSM and IDF but this means that it will not be easy to map
            the EnergyPlus results back to the original Honeybee Model. Cases
            of duplicate IDs resulting from non-unique names will be resolved
            by adding integers to the ends of the new IDs that are derived from
            the name. (Default: False).

    Returns:
        Text string of content to be written into a CSV file containing all tables
        needed to specify room loads in TRACE 700.
    """
    # get the matrices to be written to CSV format
    room_matrix, airflows_matrix, people_and_lights_matrix, misc_loads_matrix, calc_matrix = \
        model_to_trace700_matrix(
            model, si_units, ventilation_method,
            use_multiplier, exclude_plenums, merge_method, geometry_names
        )

    # put all of the matrices into one master matrix for CSV export
    mtx_width = len(room_matrix[0]) - 1
    spacer_row = ',' * mtx_width
    # add the Room table
    csv_matrix = ['ROOMS{}'.format(spacer_row)]
    for row in room_matrix:
        csv_matrix.append(','.join([str(val) for val in row]))
    # add the Airflows table
    csv_matrix.append(spacer_row)
    csv_matrix.append(spacer_row)
    csv_matrix.append('AIRFLOWS{}'.format(spacer_row))
    for row in airflows_matrix:
        csv_matrix.append(','.join([str(val) for val in row]))
    # add the People & Lighting table
    csv_matrix.append(spacer_row)
    csv_matrix.append(spacer_row)
    csv_matrix.append('PEOPLE & LIGHTING{}'.format(spacer_row))
    for row in people_and_lights_matrix:
        csv_matrix.append(','.join([str(val) for val in row]))
    # add the Miscellaneous Loads table
    csv_matrix.append(spacer_row)
    csv_matrix.append(spacer_row)
    csv_matrix.append('MISCELLANEOUS LOADS{}'.format(spacer_row))
    for row in misc_loads_matrix:
        csv_matrix.append(','.join([str(val) for val in row]))
    # add the outdoor air calculation matrix if it exists
    if calc_matrix is not None:
        csv_matrix.append(spacer_row)
        csv_matrix.append(spacer_row)
        csv_matrix.append('CALCULATION{}'.format(spacer_row))
        for row in calc_matrix:
            csv_matrix.append(','.join([str(val) for val in row]))

    return '\n'.join(csv_matrix)


def model_to_trace700_workbook(
    model, si_units=False, ventilation_method='Sum of Outdoor Air',
    use_multiplier=True, exclude_plenums=True, merge_method=None, geometry_names=False
):
    """Generate an Excel Workbook (openpyxl) with TRACE 700 attributes of a Model.

    The resulting openpyxl Workbook can be saved and opened in Excel. The data
    in the tables can then be copied into the tables that appear in the Component
    Tree view of TRACE 700. The order and organization of rooms in the resulting
    matrix should match that of the gbXML produced from the same model.

    Args:
        model: A dragonfly Model for which a TRACE 700 Excel Workbook will be returned.
        si_units: Boolean to note whether the units of the values in the resulting
            matrix are in SI (True) instead of IP (False). (Default: False).
        ventilation_method: Optional text for the ventilation method to be used in the
            resulting matrix. Choose from the following.

            * Sum of Outdoor Air
            * ASHRAE 62.1

        use_multiplier: If True, the multipliers on this Model's Stories will be
            passed along to the Workbook. If False, full geometry objects will be written
            for each and every floor in the building that are represented through
            multipliers and all resulting multipliers will be 1. (Default: True).
        exclude_plenums: Boolean to indicate whether ceiling/floor plenum depths
            assigned to Room2Ds should be ignored during translation. This
            results in each Room2D translating to a single Honeybee Room at
            the full floor_to_ceiling_height instead of a base Room with (a)
            plenum Room(s). (Default: True).
        merge_method: An optional text string to describe how the Room2Ds should
            be merged into individual Rooms during the translation. Specifying a
            value here can be an effective way to reduce the number of Room
            volumes in the resulting model and, ultimately, yield a faster
            simulation time in the destination engine with fewer results
            to manage. Note that Room2Ds will only be merged if they form a
            continuous volume. Otherwise, there will be multiple Rooms per
            zone or story, each with an integer added at the end of their
            identifiers. Choose from the following options:

            * None - No merging of Room2Ds will occur
            * Zones - Room2Ds in the same zone will be merged
            * PlenumZones - Only plenums in the same zone will be merged
            * Stories - Rooms in the same story will be merged
            * PlenumStories - Only plenums in the same story will be merged

        geometry_names: Boolean to note whether a cleaned version of all geometry
            display names should be used instead of identifiers when translating
            the Model. (Default: False).

    Returns:
        An Excel Workbook (openpyxl) with TRACE 700 attributes of the input Model.
    """
    # check that we could successfully import openpyxl
    assert openpyxl is not None, 'Export to Excel is only available in Python 3. ' \
        'Either switch to using Python 3 or use the model_to_trace700_csv instead.'

    # get the matrices to be written to CSV format
    room_matrix, airflows_matrix, people_and_lights_matrix, misc_loads_matrix, calc_matrix = \
        model_to_trace700_matrix(
            model, si_units, ventilation_method,
            use_multiplier, exclude_plenums, merge_method, geometry_names
        )

    # put all of the matrices into one master Excel workbook
    workbook = openpyxl.Workbook()
    # add the Room table
    ws = workbook.active
    _add_workbook_table(ws, 'Rooms', room_matrix)
    _apply_table_format(ws, ROOM_TABLE_FORMAT)
    # add the Airflows table
    ws_air = workbook.create_sheet('Airflows')
    _add_workbook_table(ws_air, 'Airflows', airflows_matrix)
    if ventilation_method == 'Sum of Outdoor Air':
        _apply_table_format(ws_air, AIRFLOW_TABLE_FORMAT)
    else:  # assume that it's using the TRACE native ASHRAE 62.1
        _apply_table_format(ws_air, AIRFLOW_TABLE_FORMAT_62_1)
    # add the People & Lighting table
    ws_ppl = workbook.create_sheet('People & Lighting')
    _add_workbook_table(ws_ppl, 'People & Lighting', people_and_lights_matrix)
    _apply_table_format(ws_ppl, PEOPLE_AND_LIGHTS_TABLE_FORMAT)
    # add the Miscellaneous Loads table
    ws_equip = workbook.create_sheet('Miscellaneous Loads')
    _add_workbook_table(ws_equip, 'Miscellaneous Loads', misc_loads_matrix)
    _apply_table_format(ws_equip, MISCELLANEOUS_LOADS_TABLE_FORMAT)
    # add the outdoor air calculation matrix if it exists
    if calc_matrix is not None:
        ws_oa = workbook.create_sheet('Calculation')
        _add_calc_workbook_table(ws_oa, 'Calculation', calc_matrix)
        # reference the calculated value in the airflow table
        ref_template = "='Calculation'!{}{}"
        for i, air_cell in enumerate(ws_air['7']):
            if i == 0:
                continue
            air_cell.value = ref_template.format('S', i + 3)
        for i, air_cell in enumerate(ws_air['9']):
            if i == 0:
                continue
            air_cell.value = ref_template.format('T', i + 3)
        # reference the person count in the people and lighting table
        for i, (ppl_cell, unit_cell) in enumerate(zip(ws_ppl['6'], ws_ppl['7'])):
            if i == 0:
                continue
            if unit_cell.value == 'People':
                ppl_cell.value = ref_template.format('H', i + 3)

    return workbook


def _add_workbook_table(ws, title, matrix):
    # define formatting to be used throughout the excel
    title_font = openpyxl.styles.Font(size=16, bold=True)
    bold_font = openpyxl.styles.Font(bold=True)
    side = openpyxl.styles.Side(border_style='thin', color='000000')
    all_border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    grey_fill = openpyxl.styles.PatternFill(
        start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'
    )
    row_length = len(matrix[0])
    column_letter = openpyxl.utils.get_column_letter(row_length)

    # add the title and create a border around the top row
    ws.title = title
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    for col_idx, cell in enumerate(ws['A1:{}1'.format(column_letter)][0]):
        border = cell.border
        left = side if col_idx == 0 else border.left
        right = side if col_idx == row_length - 1 else border.right
        cell.border = openpyxl.styles.Border(top=side, bottom=side, left=left, right=right)
        cell.fill = grey_fill

    # add each row of the matrix to the sheet
    for row in matrix:
        ws.append(row)
        new_row_idx = ws.max_row
        for cell in ws[new_row_idx]:
            cell.border = all_border

    # auto-fit the column width of the table to the text
    for col in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                # measure length of the cell's string representation
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except AttributeError:
                pass  # not a cell that sets the max dimension
        # apply width
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # put the first column and row in bold
    for cell in ws['A']:
        cell.font = bold_font
        cell.fill = grey_fill
    for cell in ws['2']:
        cell.font = bold_font
        cell.fill = grey_fill
    title_cell.font = title_font


def _apply_table_format(ws, formatting):
    """Apply formatting to a worksheet to make it look like a TRACE 700 table."""
    # define formatting styles to be used to make the table like TRACE
    default_font = openpyxl.styles.Font(color='FF0000')
    locked_fill = openpyxl.styles.PatternFill(
        start_color='A9A9A9', end_color='A9A9A9', fill_type='solid'
    )
    locked_font = openpyxl.styles.Font(color='808080')

    # loop through the rows and apply the formatting
    for row, tr_format in zip(ws.iter_rows(min_row=2), formatting):
        if tr_format == 'locked':
            for cell in row[1:]:
                cell.fill = locked_fill
                cell.font = locked_font
        elif tr_format == 'default':
            for cell in row[1:]:
                cell.font = default_font
        elif tr_format == 'varies':
            for cell in row[1:]:
                if isinstance(cell.value, str):
                    cell.font = default_font
                    cell.value = float(cell.value)
        elif tr_format == 'varies_default':
            for cell in row[1:]:
                if cell.value == 'Default':
                    cell.font = default_font


def _add_calc_workbook_table(ws, title, matrix):
    """Apply formatting to a the worksheet with calculations."""
    # define formatting styles to be used to make the table like TRACE
    title_font = openpyxl.styles.Font(size=16, bold=True)
    bold_font = openpyxl.styles.Font(bold=True)
    side = openpyxl.styles.Side(border_style='thin', color='000000')
    all_border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    grey_fill = openpyxl.styles.PatternFill(
        start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    text_align = openpyxl.styles.Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    row_length = len(matrix[0])
    column_letter = openpyxl.utils.get_column_letter(row_length)

    # add the title and create a border around the top row
    ws.title = title
    title_cell = ws['B1']
    title_cell.value = title
    title_cell.font = title_font
    for col_idx, cell in enumerate(ws['A1:{}1'.format(column_letter)][0]):
        border = cell.border
        left = side if col_idx == 0 else border.left
        right = side if col_idx == row_length - 1 else border.right
        cell.border = openpyxl.styles.Border(top=side, bottom=side, left=left, right=right)
        cell.fill = grey_fill

    # add each row of the matrix to the sheet
    for row in matrix:
        ws.append(row)
        new_row_idx = ws.max_row
        for cell in ws[new_row_idx]:
            cell.border = all_border

    # auto-fit the column width of the table to the text
    for i, col in enumerate(ws.columns):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col[0].column)
        if i in (0, 1, 2, 3, 4, 5, 6):
            for j, cell in enumerate(col):
                if i in (5, 6) and j == 2:
                    continue
                try:
                    # measure length of the cell's string representation
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except AttributeError:
                    pass  # not a cell that sets the max dimension
        else:
            split_str = str(col[2].value).split('\n')
            try:
                max_length = max([len(txt) for txt in split_str])
            except ValueError:  # empty cell
                pass
        # apply width
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # format the second and third columns to be readable
    for row in ws['2:3']:
        for cell in row:
            cell.font = bold_font
            cell.fill = grey_fill
            cell.alignment = text_align

    # highlight the room names and zones in gray/bold
    for i, col in enumerate(ws['A:G']):
        for j, cell in enumerate(col):
            if i == 1 and j == 0:
                continue
            cell.font = bold_font
            cell.fill = grey_fill

    # add validation to the field that permits sum or max of air
    dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"Sum,Max"')
    ws.add_data_validation(dv)
    dv.add('N4:N{}'.format(len(col)))

    # set the final air flows to be calculated from the other OA criteria
    template = '=IF(N{0}="Sum", SUM(H{0}*K{0}, I{0}*L{0}, (J{0}*M{0})/3600) / {1}{0}, ' \
        'MAX(H{0}*K{0}, I{0}*L{0}, (J{0}*M{0})/3600) / {1}{0})'
    for i, row in enumerate(ws['4:{}'.format(len(col))]):
        row_i = i + 4
        q_clg, q_htg = 'S{}'.format(row_i), 'T{}'.format(row_i)
        ws[q_clg] = template.format(row_i, 'O')
        ws[q_htg] = template.format(row_i, 'P')

    # lock and hide the column of identifiers
    unlocked = openpyxl.styles.Protection(locked=False)
    ws.protection.sheet = True
    for col in ws['B:Z']:
        for cell in col:
            cell.protection = unlocked
    ws.column_dimensions['A'].hidden = True


def model_to_trace700_gbxml(
    model, si_units=False, opening_simplification='MergeAdjWindows',
    program_name=None, program_version=None
):
    """Generate a gbXML of a model, which can be imported to TRACE 700.

    Args:
        model: A dragonfly Model for which a TRACE 700 gbXML will be returned.
        si_units: Boolean to note whether the units of the values in the resulting
            gbXML are in SI (True) instead of IP (False). (Default: False).
        opening_simplification: Optional text to note the method by which openings
            are simplified as part of the translation to gbXML. (Default: MergeAdjWindows).
            Choose from the following options.

            * None - No sub-face simplification will occur
            * MergeAdjWindows - Adjacent windows are merged; doors are left as is
            * SingleWindow - All doors removed; windows are merged into one per wall

        program_name: Optional text to set the name of the software that will
            appear under the programId and ProductName tags of the DocumentHistory
            section. This can be set things like "Ladybug Tools" or "Pollination"
            or some other software in which this gbXML export capability is being
            run. If None, the "OpenStudio" will be used. (Default: None).
        program_version: Optional text to set the version of the software that
            will appear under the DocumentHistory section. If None, and the
            program_name is also unspecified, only the version of OpenStudio will
            appear. Otherwise, this will default to "0.0.0" given that the version
            field is required. (Default: None).

    Returns:
        A gbXML text string that can be written to a file and imported to TRACE 700.
    """
    # order the room IDs so that they can be coordinated with the XLSX
    rooms_for_trace = []  # list to hold the room IDs
    assert len(model.room_2ds) != 0 or len(model.room_3ds) != 0, \
        'Model must have rooms to be able to export to TRACE700 gbXML.'
    for building in model.buildings:
        for story in building.unique_stories:
            rooms_for_trace.extend(story.room_2ds)
    for building in model.buildings:
        rooms_for_trace.extend(building.room_3ds)
    # sort the rooms so that they match their order in TRACE 700
    rooms_for_trace = sort_rooms_for_trace_700(rooms_for_trace)
    room_ids_for_trace = [r.identifier for r in rooms_for_trace]

    # set up gbXML translation parameters using the input
    gbxml_par = GBXMLParameters.for_trace_700()
    if si_units:
        gbxml_par.ip_units = False
    gbxml_par.geometry_format.opening_simplification = opening_simplification
    if str(gbxml_par.geometry_format.opening_simplification) == 'None':
        gbxml_par.geometry_format.rect_geo_format = 'SimpleAreaForNonRectOnly'
    gbxml_par.version_format.program_name = program_name
    gbxml_par.version_format.program_version = program_version

    # translate the model to a gbXML string
    return model.to_gbxml(gbxml_par, room_ids_for_trace)


def model_to_exp(model, si_units=False):
    """Get a single combined EXP string for all unique ProgramTypes and
    ConstructionSets in a Dragonfly Model.

    Args:
        model: A Dragonfly Model object.
        si_units: Boolean to note whether the units of the values in the resulting
            matrix are in SI (True) instead of IP (False). (Default: False).

    Returns:
        Text string of EXP file contents for TRACE 700.
    """
    # import the module here to avoid import failures in Python 2
    assert programs_to_exp is not None and construction_sets_to_exp is not None, \
        'Export to EXP is only available in Python 3.'

    prog_exp = programs_to_exp(model.properties.energy.program_types, si_units=si_units)
    cnst_exp = construction_sets_to_exp(
        model.properties.energy.construction_sets, si_units=si_units)

    header = 'EDITORSv6.3.1'
    prog_lines = [line for line in prog_exp.strip().splitlines() if line != header]
    cnst_lines = [line for line in cnst_exp.strip().splitlines() if line != header]

    combined_blocks = [header] + cnst_lines + prog_lines
    exp_str = '\n'.join(combined_blocks)
    return exp_str


def model_to_trace700_zip_bytes(
    model, si_units=False, opening_simplification='MergeAdjWindows',
    ventilation_method='Sum of Outdoor Air',
    program_name=None, program_version=None
):
    """Get the bytes of a .zip file containing files for import to TRACE 700.

    The .zip file will include three files within it - a gbXML that should be
    imported to TRACE 700 first to set up the geometry, an XLSX with all of
    the room properties to be pasted into the "Component View" tables of
    TRACE 700, and an optional EXP file that contains the dragonfly model's
    ProgramTypes exported to TRACE 700 room templates.

    Args:
        model: A dragonfly Model for which a TRACE 700 gbXML will be returned.
        si_units: Boolean to note whether the units of the values in the resulting
            file are in SI (True) instead of IP (False). (Default: False).
        opening_simplification: Optional text to note the method by which openings
            are simplified as part of the translation to gbXML. (Default: MergeAdjWindows).
            Choose from the following options.

            * None - No sub-face simplification will occur
            * MergeAdjWindows - Adjacent windows are merged; doors are left as is
            * SingleWindow - All doors removed; windows are merged into one per wall

        ventilation_method: Optional text for the ventilation method to be used in the
            resulting matrix. Choose from the following.

            * Sum of Outdoor Air
            * ASHRAE 62.1

        program_name: Optional text to set the name of the software that will
            appear under the programId and ProductName tags of the DocumentHistory
            section. This can be set things like "Ladybug Tools" or "Pollination"
            or some other software in which this gbXML export capability is being
            run. If None, the "OpenStudio" will be used. (Default: None).
        program_version: Optional text to set the version of the software that
            will appear under the DocumentHistory section. If None, and the
            program_name is also unspecified, only the version of OpenStudio will
            appear. Otherwise, this will default to "0.0.0" given that the version
            field is required. (Default: None).

    Returns:
        ZIP file bytes that can be written to a file and unzipped to yield everything
        needed to recreate the dragonfly model in TRACE 700.
    """
    # get the contents of the gbxml file
    gbxml_str = model_to_trace700_gbxml(
        model, si_units, opening_simplification, program_name, program_version
    )
    # get the bytes of the XLSX workbook
    xlsx_wb = model_to_trace700_workbook(model, si_units, ventilation_method)
    xlsx_stream = io.BytesIO()
    xlsx_wb.save(xlsx_stream)
    xlsx_stream.seek(0)  # reset the stream position to the beginning
    # get the contents of the EXP file
    exp_str = model_to_exp(model, si_units)

    # create an in-memory bytes buffer
    zip_buffer = io.BytesIO()
    # pass the buffer to a zip file archive
    model_id = model.display_name
    with zipfile.ZipFile(zip_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr('{}.xml'.format(model_id), gbxml_str)
        zip_file.writestr('{}.xlsx'.format(model_id), xlsx_stream.getvalue())
        zip_file.writestr('{}.exp'.format(model_id), exp_str)
    # rewind the buffer pointer to the beginning to read it
    zip_buffer.seek(0)

    # return the raw bytes of the completed ZIP file
    return zip_buffer.getvalue()
